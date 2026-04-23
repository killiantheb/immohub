"""ai_documents.py — Bail, EDL, quittances, notification, rapport, generer-document, export."""

from __future__ import annotations

import json as _json
import uuid as _uuid
from typing import Annotated

from app.core.database import get_db
from app.core.limiter import rate_limit
from app.core.security import get_current_user
from app.models.user import User
from app.services.ai_service import (
    draft_edl,
    draft_lease,
    draft_mission_report,
    draft_company_quote,
    explain_contract,
    draft_notification,
    generate_property_recap,
)
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

router = APIRouter(tags=["ai"])

DbDep       = Annotated[AsyncSession, Depends(get_db)]
AuthUserDep = Annotated[User, Depends(get_current_user)]


# ── Schemas ───────────────────────────────────────────────────────────────────

class DraftLeaseRequest(BaseModel):
    bien_id: str
    tenant_data: dict
    params: dict
    requires_validation: bool = True


class DraftEDLRequest(BaseModel):
    bien_id: str
    edl_type: str = "entry"
    inspection_date: str
    previous_edl: dict | None = None
    requires_validation: bool = True


class MissionReportRequest(BaseModel):
    mission_id: str
    observations: str


class DraftQuoteRequest(BaseModel):
    rfq_id: str
    work_description: str = ""


class ExplainContractRequest(BaseModel):
    contract_id: str


class NotificationRequest(BaseModel):
    channel: str = "email"
    recipient_role: str
    subject: str
    context: dict = {}


class PropertyRecapRequest(BaseModel):
    bien_id: str


class GenererDocumentRequest(BaseModel):
    type: str
    bien_id: _uuid.UUID
    locataire_id: _uuid.UUID | None = None
    params: dict = {}


class GenererDocumentResponse(BaseModel):
    document_id: _uuid.UUID
    url: str
    type: str


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.post("/draft-lease")
async def draft_lease_endpoint(
    payload: DraftLeaseRequest,
    current_user: AuthUserDep,
    db: DbDep,
    _=rate_limit(5, 60),
):
    """Generate a complete Swiss-law compliant lease. Owners and agencies only."""
    import uuid as _uuid_mod
    from app.models.bien import Bien
    from sqlalchemy import select as sa_sel

    if current_user.role not in ("proprio_solo", "agence", "super_admin"):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Réservé aux propriétaires et agences")

    try:
        pid = _uuid_mod.UUID(payload.bien_id)
    except ValueError:
        raise HTTPException(422, "bien_id invalide")

    result = await db.execute(sa_sel(Bien).where(Bien.id == pid))
    prop = result.scalar_one_or_none()
    if not prop:
        raise HTTPException(404, "Bien introuvable")

    property_data = {
        "type": prop.type, "address": prop.adresse, "city": prop.ville,
        "zip_code": prop.cp,
        "surface": float(prop.surface) if prop.surface else None,
        "rooms": float(prop.rooms) if prop.rooms else None,
        "floor": prop.etage, "is_furnished": prop.is_furnished,
    }

    try:
        text = await draft_lease(property_data, payload.tenant_data, payload.params, db, str(current_user.id))
    except RuntimeError as exc:
        raise HTTPException(status.HTTP_429_TOO_MANY_REQUESTS, str(exc))

    return {
        "lease_text": text,
        "requires_validation": payload.requires_validation,
        "disclaimer": "Ce bail est fourni à titre indicatif. Faites-le valider par un juriste avant signature officielle.",
    }


@router.post("/draft-edl")
async def draft_edl_endpoint(
    payload: DraftEDLRequest,
    current_user: AuthUserDep,
    db: DbDep,
    _=rate_limit(5, 60),
):
    """Generate a structured entry/exit inspection form."""
    import uuid as _uuid_mod
    from app.models.bien import Bien
    from sqlalchemy import select as sa_sel

    if current_user.role not in ("proprio_solo", "agence", "opener", "super_admin"):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Accès non autorisé")

    if payload.edl_type not in ("entry", "exit"):
        raise HTTPException(422, "edl_type doit être 'entry' ou 'exit'")

    try:
        pid = _uuid_mod.UUID(payload.bien_id)
    except ValueError:
        raise HTTPException(422, "bien_id invalide")

    result = await db.execute(sa_sel(Bien).where(Bien.id == pid))
    prop = result.scalar_one_or_none()
    if not prop:
        raise HTTPException(404, "Bien introuvable")

    property_data = {
        "type": prop.type, "address": prop.adresse, "city": prop.ville,
        "surface": float(prop.surface) if prop.surface else None,
        "rooms": float(prop.rooms) if prop.rooms else None,
        "floor": prop.etage, "is_furnished": prop.is_furnished,
        "description": prop.description_logement,
    }

    try:
        edl = await draft_edl(
            property_data, payload.edl_type, payload.inspection_date,
            payload.previous_edl, db, str(current_user.id),
        )
    except RuntimeError as exc:
        raise HTTPException(status.HTTP_429_TOO_MANY_REQUESTS, str(exc))

    return {**edl, "requires_validation": payload.requires_validation}


@router.post("/mission-report")
async def mission_report_endpoint(
    payload: MissionReportRequest,
    current_user: AuthUserDep,
    db: DbDep,
    _=rate_limit(10, 60),
):
    """Generate a professional mission report for an opener."""
    import uuid as _uuid_mod
    from app.models.opener import Mission
    from sqlalchemy import select as sa_sel

    if current_user.role not in ("opener", "super_admin"):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Réservé aux ouvreurs")

    try:
        mid = _uuid_mod.UUID(payload.mission_id)
    except ValueError:
        raise HTTPException(422, "mission_id invalide")

    result = await db.execute(sa_sel(Mission).where(Mission.id == mid))
    mission = result.scalar_one_or_none()
    if not mission:
        raise HTTPException(404, "Mission introuvable")

    mission_data = {
        "type": mission.type, "status": mission.status,
        "scheduled_at": mission.scheduled_at.isoformat() if mission.scheduled_at else None,
        "price": float(mission.price) if mission.price else None,
    }

    try:
        report = await draft_mission_report(mission_data, payload.observations, db, str(current_user.id))
    except RuntimeError as exc:
        raise HTTPException(status.HTTP_429_TOO_MANY_REQUESTS, str(exc))

    return {"report": report, "mission_id": payload.mission_id}


@router.post("/draft-quote")
async def draft_quote_endpoint(
    payload: DraftQuoteRequest,
    current_user: AuthUserDep,
    db: DbDep,
    _=rate_limit(10, 60),
):
    """AI-assisted quote draft for a company responding to an RFQ."""
    import uuid as _uuid_mod
    from app.models.rfq import RFQ
    from app.models.company import Company
    from sqlalchemy import select as sa_sel

    if current_user.role not in ("artisan", "super_admin"):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Réservé aux entreprises")

    try:
        rid = _uuid_mod.UUID(payload.rfq_id)
    except ValueError:
        raise HTTPException(422, "rfq_id invalide")

    rfq_res = await db.execute(sa_sel(RFQ).where(RFQ.id == rid))
    rfq = rfq_res.scalar_one_or_none()
    if not rfq:
        raise HTTPException(404, "Appel d'offre introuvable")

    company_res = await db.execute(sa_sel(Company).where(Company.user_id == current_user.id))
    company = company_res.scalar_one_or_none()

    rfq_data = {
        "title": rfq.title, "description": rfq.description, "category": rfq.category,
        "urgency": rfq.urgency,
        "budget_min": float(rfq.budget_min) if rfq.budget_min else None,
        "budget_max": float(rfq.budget_max) if rfq.budget_max else None,
        "city": rfq.city,
    }
    company_data = {
        "type": company.type if company else "other",
        "name": company.name if company else "",
        "rating": float(company.rating) if company and company.rating else None,
    }

    try:
        quote = await draft_company_quote(rfq_data, company_data, payload.work_description, db, str(current_user.id))
    except RuntimeError as exc:
        raise HTTPException(status.HTTP_429_TOO_MANY_REQUESTS, str(exc))

    return quote


@router.post("/explain-contract")
async def explain_contract_endpoint(
    payload: ExplainContractRequest,
    current_user: AuthUserDep,
    db: DbDep,
    _=rate_limit(10, 60),
):
    """Explain a lease contract in plain language for a tenant."""
    import uuid as _uuid_mod
    from app.models.contract import Contract
    from sqlalchemy import select as sa_sel

    try:
        cid = _uuid_mod.UUID(payload.contract_id)
    except ValueError:
        raise HTTPException(422, "contract_id invalide")

    result = await db.execute(sa_sel(Contract).where(Contract.id == cid))
    contract = result.scalar_one_or_none()
    if not contract:
        raise HTTPException(404, "Contrat introuvable")

    if current_user.role == "locataire" and contract.tenant_id != current_user.id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Accès non autorisé")

    contract_data = {
        "type": contract.type, "status": contract.status,
        "start_date": contract.start_date.isoformat() if contract.start_date else None,
        "end_date": contract.end_date.isoformat() if contract.end_date else None,
        "monthly_rent": float(contract.monthly_rent) if contract.monthly_rent else None,
        "charges": float(contract.charges) if contract.charges else None,
        "deposit": float(contract.deposit) if contract.deposit else None,
        "notice_months": getattr(contract, "notice_months", 3),
        "is_furnished": getattr(contract, "is_furnished", False),
        "pets_allowed": getattr(contract, "pets_allowed", None),
        "special_clauses": getattr(contract, "special_clauses", None),
    }

    try:
        explanation = await explain_contract(contract_data, db, str(current_user.id))
    except RuntimeError as exc:
        raise HTTPException(status.HTTP_429_TOO_MANY_REQUESTS, str(exc))

    return explanation


@router.post("/draft-notification")
async def draft_notification_endpoint(
    payload: NotificationRequest,
    current_user: AuthUserDep,
    db: DbDep,
    _=rate_limit(15, 60),
):
    """Draft a ready-to-send email or WhatsApp message."""
    if payload.channel not in ("email", "whatsapp"):
        raise HTTPException(422, "channel doit être 'email' ou 'whatsapp'")

    try:
        result = await draft_notification(
            payload.channel, payload.recipient_role, payload.subject,
            payload.context, db, str(current_user.id),
        )
    except RuntimeError as exc:
        raise HTTPException(status.HTTP_429_TOO_MANY_REQUESTS, str(exc))

    return result


@router.post("/property-recap")
async def property_recap_endpoint(
    payload: PropertyRecapRequest,
    current_user: AuthUserDep,
    db: DbDep,
    _=rate_limit(5, 60),
):
    """Generate a complete property history recap (owner/agency only)."""
    import uuid as _uuid_mod
    from app.models.bien import Bien
    from app.models.contract import Contract
    from app.models.transaction import Transaction as Txn
    from app.models.rfq import RFQ
    from sqlalchemy import select as sa_sel, and_

    if current_user.role not in ("proprio_solo", "agence", "super_admin"):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Réservé aux propriétaires et agences")

    try:
        pid = _uuid_mod.UUID(payload.bien_id)
    except ValueError:
        raise HTTPException(422, "bien_id invalide")

    result = await db.execute(sa_sel(Bien).where(Bien.id == pid))
    prop = result.scalar_one_or_none()
    if not prop:
        raise HTTPException(404, "Bien introuvable")

    contracts = (await db.execute(
        sa_sel(Contract).where(Contract.bien_id == pid)
        .order_by(Contract.start_date.desc()).limit(20)
    )).scalars().all()

    transactions = (await db.execute(
        sa_sel(Txn).where(and_(Txn.bien_id == pid, Txn.is_active.is_(True))).limit(100)
    )).scalars().all()

    rfqs = (await db.execute(
        sa_sel(RFQ).where(RFQ.bien_id == pid).order_by(RFQ.created_at.desc()).limit(20)
    )).scalars().all()

    property_data     = {"type": prop.type, "address": prop.adresse, "city": prop.ville, "status": prop.statut}
    tenants_history   = [
        {"id": str(c.id), "type": c.type, "status": c.status,
         "start_date": c.start_date.isoformat() if c.start_date else None,
         "end_date": c.end_date.isoformat() if c.end_date else None,
         "monthly_rent": float(c.monthly_rent) if c.monthly_rent else 0}
        for c in contracts
    ]
    total_revenue     = sum(float(t.amount) for t in transactions if t.status == "paid")
    unpaid            = [t for t in transactions if t.status in ("pending", "late")]
    transactions_summary = {
        "total_revenue_chf": total_revenue,
        "unpaid_count": len(unpaid),
        "unpaid_total_chf": sum(float(t.amount) for t in unpaid),
        "total_transactions": len(transactions),
    }
    interventions = [
        {"id": str(r.id), "title": r.title, "category": r.category, "status": r.status,
         "urgency": r.urgency, "created_at": r.created_at.isoformat() if r.created_at else None}
        for r in rfqs
    ]

    try:
        recap = await generate_property_recap(
            property_data, tenants_history, transactions_summary, interventions, [], db, str(current_user.id)
        )
    except RuntimeError as exc:
        raise HTTPException(status.HTTP_429_TOO_MANY_REQUESTS, str(exc))

    return recap


@router.post("/generer-document", response_model=GenererDocumentResponse)
async def generer_document(
    payload: GenererDocumentRequest,
    current_user: AuthUserDep,
    db: DbDep,
    _=rate_limit(5, 60),
) -> GenererDocumentResponse:
    """Génère un document (bail, quittance, EDL, relance) via Claude, crée un PDF et le stocke."""
    import io
    from anthropic import AsyncAnthropic
    from app.core.config import settings
    from app.models.bien import Bien
    from app.models.document_althy import DocumentAlthy
    from app.models.locataire import DossierLocataire, Locataire
    from datetime import date, datetime, timezone
    from fpdf import FPDF
    from sqlalchemy import select as sa_sel
    import httpx

    DOC_TYPES = {"bail", "quittance", "edl", "relance"}
    if payload.type not in DOC_TYPES:
        raise HTTPException(422, f"type doit être l'un de : {', '.join(DOC_TYPES)}")

    bien_res = await db.execute(sa_sel(Bien).where(Bien.id == payload.bien_id))
    bien = bien_res.scalar_one_or_none()
    if not bien:
        raise HTTPException(404, "Bien introuvable")

    loc_data: dict = {}
    if payload.locataire_id:
        loc_res = await db.execute(sa_sel(Locataire).where(Locataire.id == payload.locataire_id))
        loc = loc_res.scalar_one_or_none()
        if loc:
            loc_data = {
                "loyer": float(loc.loyer or 0), "charges": float(loc.charges or 0),
                "date_entree": loc.date_entree.isoformat() if loc.date_entree else None,
                "date_sortie": loc.date_sortie.isoformat() if loc.date_sortie else None,
            }
            dos_res = await db.execute(
                sa_sel(DossierLocataire).where(DossierLocataire.locataire_id == payload.locataire_id)
            )
            dossier = dos_res.scalar_one_or_none()
            if dossier:
                loc_data["employeur"]    = dossier.employeur
                loc_data["type_contrat"] = dossier.type_contrat

    bien_info  = f"{bien.adresse}, {bien.cp} {bien.ville} ({bien.type})"
    today_str  = date.today().strftime("%d/%m/%Y")
    params_str = _json.dumps(payload.params, ensure_ascii=False) if payload.params else "{}"

    type_prompts = {
        "bail": (
            f"Génère un bail à loyer conforme au droit suisse (CO art. 253 ss) pour le bien : {bien_info}. "
            f"Données locataire : {_json.dumps(loc_data, ensure_ascii=False)}. Paramètres : {params_str}. "
            "Structure : parties, objet, loyer/charges, durée, résiliation, dépôt, clauses spéciales."
        ),
        "quittance": (
            f"Génère une quittance de loyer pour : {bien_info}, date : {today_str}. "
            f"Données : {_json.dumps(loc_data, ensure_ascii=False)}. Params : {params_str}. "
            "Inclure : désignation du bien, montant loyer + charges, période, signature propriétaire."
        ),
        "edl": (
            "Génère un état des lieux "
            + ("d'entrée" if payload.params.get("type") == "entree" else "de sortie")
            + f" pour : {bien_info}, date : {today_str}. Params : {params_str}. "
            "Structure : pièces (entrée, séjour, cuisine, salle de bain, chambres, WC, extérieur), "
            "état de chaque élément, compteurs, clés remises. Format structuré et professionnel."
        ),
        "relance": (
            f"Rédige une lettre de relance pour loyer impayé : {bien_info}. Date : {today_str}. "
            f"Données : {_json.dumps(loc_data, ensure_ascii=False)}. Params : {params_str}. "
            "Ton professionnel mais ferme. Mentionner le montant dû, la date d'échéance, "
            "et les conséquences légales suisses en cas de non-paiement (CO art. 257d)."
        ),
    }

    try:
        client = AsyncAnthropic(api_key=settings.ANTHROPIC_API_KEY)
        response = await client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=4096,
            system="Tu es Althy, assistante immobilière suisse experte en droit du bail. "
                   "Génère des documents juridiques précis et conformes au droit suisse.",
            messages=[{"role": "user", "content": type_prompts[payload.type]}],
        )
        doc_text = response.content[0].text.strip()
    except RuntimeError as exc:
        raise HTTPException(status.HTTP_429_TOO_MANY_REQUESTS, str(exc))

    pdf = FPDF()
    pdf.set_margins(20, 20, 20)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, f"Althy — {payload.type.upper()} — {today_str}", ln=True, align="C")
    pdf.ln(5)
    pdf.set_font("Helvetica", size=10)
    for line in doc_text.split("\n"):
        pdf.multi_cell(0, 6, line if line else " ")
    pdf_bytes = pdf.output()

    ts           = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    storage_path = f"documents/{payload.type}/{payload.bien_id}/{ts}.pdf"

    async with httpx.AsyncClient(timeout=30.0) as http:
        upload_resp = await http.post(
            f"{settings.SUPABASE_URL}/storage/v1/object/althy-docs/{storage_path}",
            content=bytes(pdf_bytes),
            headers={
                "Authorization": f"Bearer {settings.SUPABASE_SERVICE_KEY}",
                "Content-Type": "application/pdf",
            },
        )
        if upload_resp.status_code not in (200, 201):
            raise HTTPException(500, f"Erreur upload Supabase: {upload_resp.text}")

    public_url = f"{settings.SUPABASE_URL}/storage/v1/object/public/althy-docs/{storage_path}"

    doc_type_map = {
        "bail": "bail",
        "quittance": "quittance",
        "edl": "edl_entree" if payload.params.get("type") != "sortie" else "edl_sortie",
        "relance": "autre",
    }
    doc = DocumentAlthy(
        bien_id=payload.bien_id, locataire_id=payload.locataire_id,
        type=doc_type_map[payload.type], url_storage=public_url,
        date_document=date.today(), genere_par_ia=True,
    )
    db.add(doc)
    await db.commit()
    await db.refresh(doc)

    return GenererDocumentResponse(document_id=doc.id, url=public_url, type=payload.type)


@router.get("/export/etat-locatif")
async def export_etat_locatif(
    current_user: AuthUserDep,
    db: DbDep,
    year: int = 2025,
):
    """Export état locatif annuel au format CSV (fiduciaire suisse, UTF-8 BOM pour Excel)."""
    import io
    import csv
    from sqlalchemy import text as sa_text

    rows = (await db.execute(
        sa_text("""
            SELECT
                to_char(p.date_echeance, 'YYYY-MM') AS mois,
                b.adresse, b.ville,
                COALESCE(u.email, 'N/A')           AS locataire,
                p.montant                            AS loyer_attendu,
                CASE WHEN p.statut = 'recu' THEN p.montant ELSE 0 END AS loyer_recu,
                COALESCE(l.charges, 0)               AS charges,
                CASE WHEN p.statut = 'recu' THEN COALESCE(p.net_montant, p.montant * 0.96) ELSE NULL END AS net,
                p.statut
            FROM paiements p
            JOIN biens b ON b.id = p.bien_id
            JOIN locataires l ON l.id = p.locataire_id
            LEFT JOIN users u ON u.id = l.user_id
            WHERE b.owner_id = :uid
              AND EXTRACT(YEAR FROM p.date_echeance) = :year
            ORDER BY p.date_echeance, b.adresse
        """),
        {"uid": str(current_user.id), "year": year},
    )).fetchall()

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow(["Mois", "Adresse", "Ville", "Locataire", "Loyer attendu CHF",
                     "Loyer reçu CHF", "Charges CHF", "Net reçu CHF", "Statut"])
    for r in rows:
        writer.writerow([
            r[0], r[1], r[2], r[3],
            f"{float(r[4] or 0):.2f}", f"{float(r[5] or 0):.2f}",
            f"{float(r[6] or 0):.2f}", f"{float(r[7] or 0):.2f}" if r[7] else "",
            r[8],
        ])
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue().encode("utf-8-sig")]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=etat_locatif_{year}.csv"},
    )


# ── SQL helper (shared by PDF/XLSX/tax exports) ──────────────────────────────

_ETAT_LOCATIF_SQL = """
    SELECT
        to_char(p.date_echeance, 'YYYY-MM') AS mois,
        b.adresse, b.ville,
        COALESCE(u.email, 'N/A')           AS locataire,
        p.montant                            AS loyer_attendu,
        CASE WHEN p.statut = 'recu' THEN p.montant ELSE 0 END AS loyer_recu,
        COALESCE(l.charges, 0)               AS charges,
        CASE WHEN p.statut = 'recu' THEN COALESCE(p.net_montant, p.montant * 0.96) ELSE NULL END AS net,
        p.statut
    FROM paiements p
    JOIN biens b ON b.id = p.bien_id
    JOIN locataires l ON l.id = p.locataire_id
    LEFT JOIN users u ON u.id = l.user_id
    WHERE b.owner_id = :uid
      AND EXTRACT(YEAR FROM p.date_echeance) = :year
    ORDER BY p.date_echeance, b.adresse
"""

_COLUMNS = ["Mois", "Adresse", "Ville", "Locataire",
            "Loyer attendu CHF", "Loyer reçu CHF", "Charges CHF", "Net reçu CHF", "Statut"]


@router.get("/export/etat-locatif-pdf")
async def export_etat_locatif_pdf(
    current_user: AuthUserDep,
    db: DbDep,
    year: int = 2025,
):
    """Export état locatif annuel au format PDF (fpdf2)."""
    import io
    from fpdf import FPDF
    from sqlalchemy import text as sa_text

    rows = (await db.execute(
        sa_text(_ETAT_LOCATIF_SQL), {"uid": str(current_user.id), "year": year},
    )).fetchall()

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, f"Etat locatif annuel {year}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 6, f"Genere par Althy · {current_user.email}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    col_w = [22, 50, 30, 50, 28, 28, 22, 28, 20]
    pdf.set_font("Helvetica", "B", 7.5)
    for i, h in enumerate(_COLUMNS):
        pdf.cell(col_w[i], 7, h, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", "", 7.5)
    for r in rows:
        vals = [
            str(r[0] or ""),
            str(r[1] or "")[:28],
            str(r[2] or ""),
            str(r[3] or "")[:28],
            f"{float(r[4] or 0):.2f}",
            f"{float(r[5] or 0):.2f}",
            f"{float(r[6] or 0):.2f}",
            f"{float(r[7] or 0):.2f}" if r[7] else "",
            str(r[8] or ""),
        ]
        for i, v in enumerate(vals):
            pdf.cell(col_w[i], 6, v, border=1)
        pdf.ln()

    buf = io.BytesIO(pdf.output())
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=etat_locatif_{year}.pdf"},
    )


@router.get("/export/etat-locatif-xlsx")
async def export_etat_locatif_xlsx(
    current_user: AuthUserDep,
    db: DbDep,
    year: int = 2025,
):
    """Export état locatif annuel au format XLSX (openpyxl), compatible ERP suisse."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from sqlalchemy import text as sa_text

    rows = (await db.execute(
        sa_text(_ETAT_LOCATIF_SQL), {"uid": str(current_user.id), "year": year},
    )).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Etat locatif {year}"

    # Header row
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="E8602C", end_color="E8602C", fill_type="solid")
    header_text = Font(bold=True, size=10, color="FFFFFF")
    for c, h in enumerate(_COLUMNS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_text
        cell.fill = header_fill

    for r_idx, r in enumerate(rows, 2):
        ws.cell(row=r_idx, column=1, value=str(r[0] or ""))
        ws.cell(row=r_idx, column=2, value=str(r[1] or ""))
        ws.cell(row=r_idx, column=3, value=str(r[2] or ""))
        ws.cell(row=r_idx, column=4, value=str(r[3] or ""))
        ws.cell(row=r_idx, column=5, value=float(r[4] or 0))
        ws.cell(row=r_idx, column=6, value=float(r[5] or 0))
        ws.cell(row=r_idx, column=7, value=float(r[6] or 0))
        ws.cell(row=r_idx, column=8, value=float(r[7] or 0) if r[7] else None)
        ws.cell(row=r_idx, column=9, value=str(r[8] or ""))

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=etat_locatif_{year}.xlsx"},
    )


@router.get("/export/declaration-fiscale")
async def export_declaration_fiscale(
    current_user: AuthUserDep,
    db: DbDep,
    year: int = 2025,
):
    """Génère un PDF de déclaration fiscale préremplie (revenus locatifs)."""
    import io
    from fpdf import FPDF
    from sqlalchemy import text as sa_text

    # Aggregate per-property
    rows = (await db.execute(
        sa_text("""
            SELECT
                b.adresse, b.ville,
                SUM(CASE WHEN p.statut = 'recu' THEN p.montant ELSE 0 END) AS total_recu,
                SUM(p.montant)                                              AS total_attendu,
                SUM(COALESCE(l.charges, 0))                                 AS total_charges
            FROM paiements p
            JOIN biens b ON b.id = p.bien_id
            JOIN locataires l ON l.id = p.locataire_id
            WHERE b.owner_id = :uid
              AND EXTRACT(YEAR FROM p.date_echeance) = :year
            GROUP BY b.adresse, b.ville
            ORDER BY b.adresse
        """),
        {"uid": str(current_user.id), "year": year},
    )).fetchall()

    total_brut = sum(float(r[2] or 0) for r in rows)
    total_charges = sum(float(r[4] or 0) for r in rows)
    total_net = total_brut - total_charges

    pdf = FPDF(unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, f"Declaration fiscale — Revenus locatifs {year}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, f"Contribuable : {current_user.first_name or ''} {current_user.last_name or ''} — {current_user.email}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, "Document genere par Althy — a joindre a votre declaration d'impot", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)

    # Per-property table
    pdf.set_font("Helvetica", "B", 9)
    col_w = [60, 30, 35, 35, 30]
    headers = ["Bien", "Ville", "Loyers bruts CHF", "Charges CHF", "Net CHF"]
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 7, h, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", "", 9)
    for r in rows:
        brut = float(r[2] or 0)
        charges = float(r[4] or 0)
        pdf.cell(col_w[0], 6, str(r[0] or "")[:32], border=1)
        pdf.cell(col_w[1], 6, str(r[1] or ""), border=1)
        pdf.cell(col_w[2], 6, f"{brut:,.2f}", border=1)
        pdf.cell(col_w[3], 6, f"{charges:,.2f}", border=1)
        pdf.cell(col_w[4], 6, f"{brut - charges:,.2f}", border=1)
        pdf.ln()

    # Totals
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(col_w[0] + col_w[1], 7, "TOTAL", border=1)
    pdf.cell(col_w[2], 7, f"{total_brut:,.2f}", border=1)
    pdf.cell(col_w[3], 7, f"{total_charges:,.2f}", border=1)
    pdf.cell(col_w[4], 7, f"{total_net:,.2f}", border=1)
    pdf.ln(14)

    pdf.set_font("Helvetica", "", 8)
    pdf.multi_cell(0, 5, (
        "Ce document est un recapitulatif genere automatiquement par Althy. "
        "Il ne constitue pas un formulaire fiscal officiel. "
        "Veuillez reporter ces montants dans votre declaration d'impot "
        "(formulaire cantonal, annexe immeubles)."
    ))

    buf = io.BytesIO(pdf.output())
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=declaration_fiscale_{year}.pdf"},
    )


@router.get("/export/rapport-gestion")
async def export_rapport_gestion(
    current_user: AuthUserDep,
    db: DbDep,
    year: int = 2025,
):
    """Rapport de gestion annuel — performances et rendements par bien."""
    import io
    from fpdf import FPDF
    from sqlalchemy import text as sa_text

    rows = (await db.execute(
        sa_text("""
            SELECT
                b.adresse, b.ville,
                COUNT(DISTINCT l.id)                                        AS nb_locataires,
                SUM(p.montant)                                              AS total_attendu,
                SUM(CASE WHEN p.statut = 'recu' THEN p.montant ELSE 0 END) AS total_recu,
                SUM(CASE WHEN p.statut = 'en_retard' THEN p.montant ELSE 0 END) AS total_impayes
            FROM paiements p
            JOIN biens b ON b.id = p.bien_id
            JOIN locataires l ON l.id = p.locataire_id
            WHERE b.owner_id = :uid
              AND EXTRACT(YEAR FROM p.date_echeance) = :year
            GROUP BY b.adresse, b.ville
            ORDER BY b.adresse
        """),
        {"uid": str(current_user.id), "year": year},
    )).fetchall()

    total_attendu = sum(float(r[3] or 0) for r in rows)
    total_recu = sum(float(r[4] or 0) for r in rows)
    total_impayes = sum(float(r[5] or 0) for r in rows)
    taux_recouvrement = round(total_recu / total_attendu * 100, 1) if total_attendu else 0

    pdf = FPDF(unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, f"Rapport de gestion {year}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, f"Proprietaire : {current_user.first_name or ''} {current_user.last_name or ''}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"Genere par Althy le {__import__('datetime').date.today().isoformat()}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)

    # KPIs
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 7, "Indicateurs cles", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, f"  Loyers attendus :     CHF {total_attendu:,.2f}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"  Loyers encaisses :    CHF {total_recu:,.2f}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"  Impayes :             CHF {total_impayes:,.2f}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"  Taux de recouvrement: {taux_recouvrement}%", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"  Nombre de biens :     {len(rows)}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)

    # Per-property table
    pdf.set_font("Helvetica", "B", 9)
    col_w = [50, 25, 20, 30, 30, 30]
    headers = ["Bien", "Ville", "Locataires", "Attendu CHF", "Recu CHF", "Impayes CHF"]
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 7, h, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", "", 9)
    for r in rows:
        pdf.cell(col_w[0], 6, str(r[0] or "")[:28], border=1)
        pdf.cell(col_w[1], 6, str(r[1] or ""), border=1)
        pdf.cell(col_w[2], 6, str(int(r[2] or 0)), border=1)
        pdf.cell(col_w[3], 6, f"{float(r[3] or 0):,.2f}", border=1)
        pdf.cell(col_w[4], 6, f"{float(r[4] or 0):,.2f}", border=1)
        pdf.cell(col_w[5], 6, f"{float(r[5] or 0):,.2f}", border=1)
        pdf.ln()

    buf = io.BytesIO(pdf.output())
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=rapport_gestion_{year}.pdf"},
    )
